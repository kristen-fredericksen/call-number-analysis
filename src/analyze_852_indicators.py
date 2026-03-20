#!/usr/bin/env python3
"""
852 First Indicator Analysis Script

Analyzes MARC 852 fields to suggest correct first indicator values based on
call number content. Developed for CUNY consortium data but applicable to
any library using LC, Dewey, SuDoc, NLM, or local classification schemes.

Usage:
    python analyze_852_indicators.py input.xlsx              # auto-generates timestamped output
    python analyze_852_indicators.py input.xlsx output.xlsx  # custom output name

Input file should have columns:
    - Permanent Call Number
    - Permanent Call Number Type
    - 852 MARC
    - Normalized Call Number
    - Institution Name
    - MMS Id

Output includes suggested indicators, classification types, confidence levels,
and notes explaining the classification decision.

Author: Developed iteratively with Claude
Version: 1.0
"""

import pandas as pd
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# =============================================================================
# CUNY INSTITUTION CODES (for Primo source record links)
# =============================================================================

# Maps Alma IZ institution codes to institution names.
# Analytics sometimes returns the IZ code instead of the institution name.
ALMA_IZ_CODES = {
    '6121': 'CUNY Network Zone',
    '6122': 'Baruch College',
    '6123': 'Bronx Community College',
    '6124': 'Brooklyn College',
    '6125': 'College of Staten Island',
    '6126': 'Craig Newmark Graduate School of Journalism at CUNY',
    '6127': 'LaGuardia Community College',
    '6128': 'John Jay College of Criminal Justice',
    '6129': 'Kingsborough Community College',
    '6130': 'CUNY School of Law',
    '6131': 'Guttman Community College',
    '6132': 'Hostos Community College',
    '6133': 'Hunter College',
    '6134': 'Lehman College',
    '6135': 'Medgar Evers College',
    '6136': 'New York City College of Technology',
    '6137': 'Queens College',
    '6138': 'The City College of New York',
    '6139': 'York College',
    '6140': 'CUNY Graduate Center',
    '6141': 'Borough of Manhattan Community College',
    '6142': 'Queensborough Community College',
    '6143': 'CUNY Central Office',
}

# Maps institution names (as they appear in Analytics) to CUNY Primo codes.
# Used to build permalink URLs to Primo VE.
CUNY_PRIMO_CODES = {
    'Baruch College': 'BB',
    'Borough of Manhattan Community College': 'BM',
    'Bronx Community College': 'BX',
    'Brooklyn College': 'BC',
    'College of Staten Island': 'SI',
    'Craig Newmark Graduate School of Journalism at CUNY': 'GJ',
    'CUNY Central Office': 'AL',
    'CUNY Graduate Center': 'GC',
    'CUNY School of Law': 'CL',
    'Guttman Community College': 'NC',
    'Hostos Community College': 'HO',
    'Hunter College': 'HC',
    'John Jay College of Criminal Justice': 'JJ',
    'Kingsborough Community College': 'KB',
    'LaGuardia Community College': 'LG',
    'Lehman College': 'LE',
    'Medgar Evers College': 'ME',
    'New York City College of Technology': 'NY',
    'Queens College': 'QC',
    'Queensborough Community College': 'QB',
    'The City College of New York': 'CC',
    'York College': 'YC',
}

# Primo VE search profile scope codes (IZ scope) per school code.
# Used in permalink URLs:
#   https://cuny-kb.primo.exlibrisgroup.com/permalink/01CUNY_KB/169j6ul/alma{MMS_ID}
# Add scope codes here as schools are onboarded.
PRIMO_SCOPE_CODES = {
    'BM': 'pafot',
    'KB': '60uq85',
    'QC': '1j7jakf',
}


# =============================================================================
# MARC 852 PARSING
# =============================================================================

def parse_852_marc(marc_field):
    """
    Parse 852 MARC field into components.
    
    Example input: 852_0 $$a NBC $$b BC001 $$c FOLIO $$h N620 .F6 $$i A85 $$k FOLIO
    
    Returns dict with indicators and subfields, or None if input is empty.
    """
    if pd.isna(marc_field):
        return None
    
    marc = str(marc_field)
    result = {'raw': marc, 'subfields': {}}
    
    # Extract indicators (852_0, 852__, 852#4, etc.)
    ind_match = re.match(r'^852([_#0-9])([_#0-9])', marc)
    if ind_match:
        ind1, ind2 = ind_match.groups()
        result['indicator1'] = ind1 if ind1 not in ['_', '#'] else ''
        result['indicator2'] = ind2 if ind2 not in ['_', '#'] else ''
    
    # Extract subfields ($$a, $$b, $$h, $$i, $$k, etc.)
    subfield_pattern = r'\$\$([a-z0-9])\s*([^$]*)'
    for match in re.finditer(subfield_pattern, marc, re.IGNORECASE):
        code = match.group(1).lower()
        value = match.group(2).strip()
        if code in result['subfields']:
            result['subfields'][code] += ' ' + value
        else:
            result['subfields'][code] = value
    
    return result


def _looks_like_shelving_control(value):
    """
    Check if a $j value looks like a shelving control number
    (format + accession number) rather than a miscoded cutter.

    Examples that ARE shelving control: DVD 521, CD 1811, Video disc 110
    Examples that are NOT (miscoded cutters): .A85, M53, R74
    """
    if _RE_SHELVING_CTRL_AV.match(value.strip()):
        return True
    return False


def get_call_number_from_marc(parsed_marc):
    """
    Extract the actual call number from parsed 852, using $$h and $$i (or $$j).

    IMPORTANT: Ignores $$k (prefix) for classification purposes.
    The prefix (FOLIO, REF, OVERSIZE, etc.) should not affect scheme identification.

    Subfield priority:
    - $$j (shelving control number) - used for indicator 4
    - $$h + $$i (classification + item) - used for LC, Dewey, etc.

    Returns: (call_number, from_j, j_combined, j_conflict) where:
    - from_j: True if value came from $$j alone
    - j_combined: True if $$j was merged with $$h/$$i (miscoded cutter)
    - j_conflict: True if $$h has a classification AND $$j has a separate
      shelving control number (conflicting schemes in one 852 field)
    """
    if not parsed_marc:
        return None, False, False, False

    subfields = parsed_marc.get('subfields', {})

    h = subfields.get('h', '')
    i = subfields.get('i', '')
    j = subfields.get('j', '')

    # Strip trailing period from $h for classification purposes.
    # A trailing period on $h (e.g., "PL955.") usually means the period
    # belongs at the start of the cutter in $i (e.g., ".K4913").
    # Leaving it in prevents the classifier from recognizing the pattern.
    if h and h.endswith('.') and i:
        h = h[:-1].rstrip()

    if j:
        if h or i:
            if _looks_like_shelving_control(j):
                # $h has a classification and $j has a separate shelving
                # control number — two schemes in one field. Classify
                # from $h/$i only; flag the conflict.
                if i:
                    return f"{h} {i}".strip(), False, False, True
                return h.strip(), False, False, True
            # $j alongside $h/$i but $j doesn't look like a shelving
            # control number — likely a miscoded cutter.
            parts = [p for p in [h, i, j] if p]
            return ' '.join(parts).strip(), False, True, False
        # $j only (no $h or $i) — standalone shelving control number
        return j.strip(), True, False, False
    elif h:
        if i:
            return f"{h} {i}".strip(), False, False, False
        return h.strip(), False, False, False
    elif i:
        # $i (item part/cutter) without $h (classification) — the class
        # number is missing. Return $i so it's visible in the output, but
        # it will be flagged during classification.
        return i.strip(), False, False, False

    return None, False, False, False


# =============================================================================
# VALID LC CLASSIFICATION LETTERS
# =============================================================================

# Complete list of LC classification letters, extracted from the Library of
# Congress Classification schedule PDFs (A through Z), downloaded Feb 2026.
# LC uses A-Z EXCEPT I, O, W, X, Y:
# - I and O are not used (avoid confusion with 1 and 0)
# - W is NLM (medicine)
# - X is not used
# - Y is SuDoc (Congressional)

LC_VALID_CLASSES = {
    # A - General Works
    'A', 'AC', 'AE', 'AG', 'AI', 'AM', 'AN', 'AP', 'AS', 'AY', 'AZ',
    # B - Philosophy, Psychology, Religion
    'B', 'BC', 'BD', 'BF', 'BH', 'BJ', 'BL', 'BM', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BV', 'BX',
    # C - Auxiliary Sciences of History
    'C', 'CB', 'CC', 'CD', 'CE', 'CJ', 'CN', 'CR', 'CS', 'CT',
    # D - World History
    'D', 'DA', 'DAW', 'DB', 'DC', 'DD', 'DE', 'DF', 'DG', 'DH', 'DJ', 'DJK',
    'DK', 'DL', 'DP', 'DQ', 'DR', 'DS', 'DT', 'DU', 'DX',
    # E, F - History of the Americas (no subclasses)
    'E', 'F',
    # G - Geography, Anthropology, Recreation
    'G', 'GA', 'GB', 'GC', 'GE', 'GF', 'GN', 'GR', 'GT', 'GV',
    # H - Social Sciences
    'H', 'HA', 'HB', 'HC', 'HD', 'HE', 'HF', 'HG', 'HJ', 'HM', 'HN', 'HQ', 'HS', 'HT', 'HV', 'HX',
    # J - Political Science
    'J', 'JA', 'JC', 'JF', 'JJ', 'JK', 'JL', 'JN', 'JQ', 'JS', 'JV', 'JX', 'JZ',
    # K - Law (two-letter subclasses)
    'K', 'KB', 'KD', 'KE', 'KF', 'KG', 'KH', 'KJ', 'KK', 'KL', 'KM', 'KN', 'KP', 'KQ', 'KR', 'KS', 'KT', 'KU', 'KV', 'KW', 'KZ',
    # K - Law (three-letter subclasses)
    'KBM', 'KBP', 'KBR', 'KBU',
    'KDC', 'KDE', 'KDG', 'KDK', 'KDZ',
    'KEA', 'KEB', 'KEM', 'KEN', 'KEO', 'KEP', 'KEQ', 'KES', 'KEY', 'KEZ',
    'KFA', 'KFC', 'KFD', 'KFF', 'KFG', 'KFH', 'KFI', 'KFK', 'KFL', 'KFM',
    'KFN', 'KFO', 'KFP', 'KFR', 'KFS', 'KFT', 'KFU', 'KFV', 'KFW', 'KFX', 'KFZ',
    'KGA', 'KGB', 'KGC', 'KGD', 'KGE', 'KGF', 'KGG', 'KGH', 'KGJ', 'KGK',
    'KGL', 'KGM', 'KGN', 'KGP', 'KGQ', 'KGR', 'KGS', 'KGT', 'KGU', 'KGV',
    'KGW', 'KGX', 'KGY', 'KGZ',
    'KHA', 'KHC', 'KHD', 'KHF', 'KHH', 'KHK', 'KHL', 'KHM', 'KHN', 'KHP',
    'KHQ', 'KHS', 'KHU', 'KHW',
    'KJA', 'KJC', 'KJE', 'KJG', 'KJH', 'KJJ', 'KJK', 'KJM', 'KJN',
    'KJP', 'KJR', 'KJS', 'KJT', 'KJV', 'KJW',
    'KKA', 'KKB', 'KKC', 'KKE', 'KKF', 'KKG', 'KKH', 'KKI', 'KKJ',
    'KKK', 'KKL', 'KKM', 'KKN', 'KKP', 'KKQ', 'KKR', 'KKS', 'KKT', 'KKV',
    'KKW', 'KKX', 'KKY', 'KKZ',
    'KLA', 'KLB', 'KLD', 'KLE', 'KLF', 'KLH', 'KLM', 'KLN', 'KLP',
    'KLQ', 'KLR', 'KLS', 'KLT', 'KLV', 'KLW',
    'KMC', 'KME', 'KMF', 'KMG', 'KMH', 'KMJ', 'KMK', 'KML', 'KMM',
    'KMN', 'KMP', 'KMQ', 'KMS', 'KMT', 'KMU', 'KMV', 'KMX', 'KMY',
    'KNC', 'KNE', 'KNF', 'KNG', 'KNH', 'KNK', 'KNL', 'KNM', 'KNN',
    'KNP', 'KNQ', 'KNR', 'KNS', 'KNT', 'KNU', 'KNV', 'KNW', 'KNX', 'KNY',
    'KPA', 'KPC', 'KPE', 'KPF', 'KPG', 'KPH', 'KPJ', 'KPK', 'KPL',
    'KPM', 'KPP', 'KPS', 'KPT', 'KPV', 'KPW',
    'KQC', 'KQE', 'KQG', 'KQH', 'KQJ', 'KQK', 'KQM', 'KQP', 'KQT',
    'KQV', 'KQW', 'KQX',
    'KRB', 'KRC', 'KRE', 'KRG', 'KRK', 'KRL', 'KRM', 'KRN', 'KRP',
    'KRR', 'KRS', 'KRU', 'KRV', 'KRW', 'KRX', 'KRY',
    'KSA', 'KSC', 'KSE', 'KSG', 'KSH', 'KSK', 'KSL', 'KSN', 'KSP',
    'KSR', 'KSS', 'KST', 'KSU', 'KSV', 'KSW', 'KSX', 'KSY', 'KSZ',
    'KTA', 'KTC', 'KTD', 'KTE', 'KTF', 'KTG', 'KTH', 'KTJ', 'KTK',
    'KTL', 'KTN', 'KTQ', 'KTR', 'KTT', 'KTU', 'KTV', 'KTW', 'KTX', 'KTY', 'KTZ',
    'KUA', 'KUN', 'KUQ',
    'KVB', 'KVC', 'KVE', 'KVH', 'KVL', 'KVM', 'KVN', 'KVP', 'KVQ',
    'KVR', 'KVS', 'KVU', 'KVW',
    'KWA', 'KWC', 'KWE', 'KWG', 'KWH', 'KWL', 'KWP', 'KWQ', 'KWR',
    'KWT', 'KWW', 'KWX',
    'KZA', 'KZD',
    # L - Education
    'L', 'LA', 'LB', 'LC', 'LD', 'LE', 'LF', 'LG', 'LH', 'LJ', 'LT',
    # M - Music
    'M', 'ML', 'MT',
    # N - Fine Arts
    'N', 'NA', 'NB', 'NC', 'ND', 'NE', 'NK', 'NX',
    # P - Language and Literature
    'P', 'PA', 'PB', 'PC', 'PD', 'PE', 'PF', 'PG', 'PH', 'PJ', 'PK', 'PL', 'PM',
    'PN', 'PQ', 'PR', 'PS', 'PT', 'PZ',
    # Q - Science
    'Q', 'QA', 'QB', 'QC', 'QD', 'QE', 'QH', 'QK', 'QL', 'QM', 'QP', 'QR',
    # R - Medicine
    'R', 'RA', 'RB', 'RC', 'RD', 'RE', 'RF', 'RG', 'RJ', 'RK', 'RL', 'RM', 'RS', 'RT', 'RV', 'RX', 'RZ',
    # S - Agriculture
    'S', 'SB', 'SD', 'SF', 'SH', 'SK',
    # T - Technology
    'T', 'TA', 'TC', 'TD', 'TE', 'TF', 'TG', 'TH', 'TJ', 'TK', 'TL', 'TN', 'TP', 'TR', 'TS', 'TT', 'TX',
    # U - Military Science
    'U', 'UA', 'UB', 'UC', 'UD', 'UE', 'UF', 'UG', 'UH',
    # V - Naval Science
    'V', 'VA', 'VB', 'VC', 'VD', 'VE', 'VF', 'VG', 'VK', 'VM',
    # Z - Bibliography, Library Science
    'Z', 'ZA',
}


# =============================================================================
# CLASSIFICATION DETECTION FUNCTIONS
# =============================================================================

# Exact match values for non-call-number detection (computed once)
_TEST_VALUES = {
    'test', 'sample', 'example', 'dummy', 'temp', 'temporary',
    'xxx', 'zzz', 'tbd', 'tba', 'n/a', 'na', 'none', 'null',
    'delete', 'remove', 'fix', 'error', 'blank', 'empty'
}

_FORMAT_ONLY = {
    'cd rom', 'cd-rom', 'cdrom', 'dvd rom', 'dvd-rom', 'dvdrom',
    'dvd', 'cd', 'vhs', 'dvd video', 'computer file',
    'cassette', 'microfilm', 'microfiche', 'filmstrip'
}

_EQUIPMENT_WORDS = {
    'projector', 'marker', 'charger', 'adapter', 'cable', 'remote',
    'headphones', 'headset', 'speaker', 'tripod', 'screen', 'pointer',
    'clicker', 'eraser', 'whiteboard', 'easel', 'calculator',
}

# Public note patterns — patron-facing instructions that belong in $z
_PUBLIC_NOTE_PATTERNS = [re.compile(p) for p in [
    # Patron instructions
    r'(?i)ask\s+(at|for|librarian|staff)',
    r'(?i)check\s+(with|at|below)',
    r'(?i)inquire',
    r'(?i)please\s+',
    r'(?i)assistance',
    r'(?i)circulation\s+desk',
    r'(?i)microforms?\s+desk',

    # Access/availability notes (patron-facing)
    r'(?i)^access\s*(for|through|to|:)',
    r'(?i)access[:.]',
    r'(?i)available\s+(at|in|from|to|on)',
    r'(?i)not\s+available',
    r'(?i)users?\s+(only|must|can|may)',
    r'(?i)(college|university|library)\s+users',

    # Shelving/location guidance for patrons
    r'(?i)shelved\s+(with|in|at|by|under)',
    r'(?i)filed\s+(under|with|in)',
    r'(?i)located\s+(in|at)',
    r'(?i)\bon\s+reserve',
    r'(?i)reference\s+(only|desk|room)',
    r'(?i)non-circulating',
    r'(?i)in-library\s+use',
    r'(?i)room\s+use\s+only',
    r'(?i)does\s+not\s+circulate',
    r'(?i)\brestricted\b',
    r'(?i)permission\s+(required|needed)',

    # Electronic access
    r'(?i)online\s+(access|only|version)',
    r'(?i)electronic\s+(access|version)',
    r'(?i)e-?resource',
    r'(?i)\bdatabase\b',
    r'(?i)website',
    r'(?i)workstation',
    r'(?i)^https?://',
    r'(?i)^www\.',

    # Browsing notes
    r'(?i)current\s+issues',

    # Location/room references
    r'(?i)reading\s+room',
    r'(?i)\d+(st|nd|rd|th)\s+floor',

    # Format + location notes (e.g., "Bound volumes: 3rd floor")
    r'(?i)bound\s+volume',

    # Loan period/circulation notes
    r'(?i)\d+[- ]?(day|week|hour)\s+(loan|checkout|reserve)',
]]

# Staff note patterns — cataloging/processing notes that belong in $x
_STAFF_NOTE_PATTERNS = [re.compile(p) for p in [
    # Cataloging notes
    r'(?i)cataloged\s+(under|with|as|separately)',
    r'(?i)classed\s+(with|in)',
    r'(?i)search\s+under',
    r'(?i)see\s+(also|librarian|reference|archivist)',
    r'(?i)contact\s+',
    r'(?i)request\s+(from|at|through)',
    r'(?i)consult\s+',
    r'(?i)bound\s+with',
    r'(?i)use\s+copy\s+in',
    r'(?i)keep\s+at',
    r'(?i)kept\s+(on|at|in)',
    r'(?i)stored\s+(off-?site|in)',
    r'(?i)shelve\s+in\b',
    r'(?i)library\s+copy',
    r'(?i)scholarship',
    r'(?i)order\s+(from|through)',
    r'(?i)interlibrary\s+loan',
    r'(?i)\bill\s+only',

    # Status notes
    r'(?i)^in\s+process',
    r'(?i)superseded',
    r'(?i)cancelled',
    r'(?i)withdrawn',
    r'(?i)\bmissing\b',
    r'(?i)\blost\b',
    r'(?i)damaged',

    # Shelving instructions (staff-facing)
    r'(?i)sort\s+by\s+',
    r'(?i)separately\s+classed',
    r'(?i)shelved\s+alphabetically',

    # Volume/issue notation without call number
    r'(?i)^\*\s*(vol|no\.?|v\.|issue|pt\.?|part)',

    # Encoding/placeholder patterns
    r'(?i)^e-[a-z]{2}---',
    r'(?i)^[a-z]-[a-z]{2}---',
]]

_PUNCTUATION_ONLY_RE = re.compile(r'^[\?\.\-\_\*\#]+$')

# Pre-compiled classification patterns (compiled once at import, not per call)
_RE_DIGIT = re.compile(r'\d')

# AV shelving detection
_RE_SHELVING_CTRL_AV = re.compile(r'^(DVD|CD|VHS|Video|Fiche|Disc|Tape|Cassette)\b', re.IGNORECASE)
_RE_AV_SIMPLE = re.compile(r'^(CD|DVD|VHS|LP|MC|DAT)[\s\-]+[A-Z]*\d+(\s|$)', re.IGNORECASE)
_RE_AV_LC_CHECK = re.compile(r'^(CD|DVD)\s*\d+[\s.]+\.?[A-Z]\d*', re.IGNORECASE)
_RE_AV_ROM = re.compile(r'^(CD|DVD)[\s\-]*(ROM)\s+\d+', re.IGNORECASE)
_RE_AV_PREFIX_FORMAT = re.compile(r'^[A-Z]+\s+(CD|DVD)[\s\-]*(ROM)?\s+\d+', re.IGNORECASE)
_RE_AV_VIDEO = re.compile(r'Video[\s\-]*(disc|recording|tape|CD|DVD|VHS)\s*[A-Z]*\d+', re.IGNORECASE)
_RE_AV_CASSETTE = re.compile(r'^VIDEO\s+CASSETTE\s+\d+', re.IGNORECASE)
_RE_AV_FICHE = re.compile(r'^(Fiche|Micro(film|card|fiche))\s*\d+', re.IGNORECASE)
_RE_AV_MICRO_FORMAT = re.compile(r'^Micro(film|card|fiche)\s+[A-Z]+\s+\d+', re.IGNORECASE)
_RE_AV_RECORDING = re.compile(r'Recording\s+[A-Z]*\d+', re.IGNORECASE)
_RE_AV_MUSIC = re.compile(r'^Music\s+(CD|DVD)\s+', re.IGNORECASE)

# Local collection schemes
_RE_LOCAL_HYPHEN = re.compile(r'^[A-Z]{2,5}\s+\d{2,4}-\d+', re.IGNORECASE)
_RE_LOCAL_PREFIX_NUM = re.compile(r'^[A-Z]{2,5}\s+\d{2,4}(\s|$)', re.IGNORECASE)

# Class letter extraction
_RE_CLASS_LETTERS = re.compile(r'^([A-Z]{1,3})\s*\d', re.IGNORECASE)

# SuDoc
_RE_SUDOC = re.compile(r'^[A-Z]{1,4}\s*\d+\.[A-Z0-9\s/\-\.]+:', re.IGNORECASE)
_RE_SUDOC_Y = re.compile(r'^Y\s*\d', re.IGNORECASE)

# LC geographic cutter colon — colons in LC cutter notation for subordinate
# locations (Tables G1548-G9804). Pattern: .Letter+Digits:Digits+Letter+Digits
# Examples: G3424 .A35:2C3 (Canadian Forces Base at Aldershot, Nova Scotia),
#           G1778 .P4:3C8 (Curitiba, Brazil)
_RE_LC_CUTTER_COLON = re.compile(
    r'^[A-Z]{1,3}\s*\d+\s*\.[A-Z]\d+:\d+[A-Z]\d+',
    re.IGNORECASE
)

# NLM
_RE_NLM_CLASS = re.compile(r'^(Q[S-Z]|W[A-Z]?)$')
_RE_NLM_NUM = re.compile(r'^(Q[S-Z]|W[A-Z]?)\s*\d')

# LAC (Library and Archives Canada)
_RE_LAC_FC = re.compile(r'^FC\s*\d', re.IGNORECASE)
_RE_LAC_PS = re.compile(r'^PS\s*(\d+)', re.IGNORECASE)

# Dewey
_RE_DEWEY_DECIMAL = re.compile(r'^\d{3}\.\d+')
_RE_DEWEY_CUTTER = re.compile(r'^(\d{3})\s+([A-Z]\d+[A-Z]?)(\s|$)', re.IGNORECASE)
_RE_DEWEY_AUTHOR = re.compile(r'^(\d{3})\s+[A-Z][a-z]{1,4}\b')

# Local reserve labels
_RE_RESERVE_EDITION = re.compile(r'^[A-Za-z]+\s*\d{4}\s+\d+(st|nd|rd|th)\s+Ed', re.IGNORECASE)
_RE_RESERVE_YEAR = re.compile(r'^[A-Z]{1,3}\s+\d{1,2}\s+\d{4}\s*$', re.IGNORECASE)

# LC classification
_RE_LC_MLCS = re.compile(r'MLCS\s*\d{4}/', re.IGNORECASE)
_RE_LC_CUTTER_ATTACHED = re.compile(r'^[A-Z]{1,3}\s*\d{1,4}\.[A-Z]\d*', re.IGNORECASE)
_RE_LC_CUTTER_SPACE = re.compile(r'^[A-Z]{1,3}\s*\d{1,4}(\s*\.\d+)?\s+\.?[A-Z]\d*', re.IGNORECASE)
_RE_LC_DATE_CUTTER = re.compile(r'^[A-Z]{1,3}\s*\d{1,4}\s+\d{4}\s+\.?[A-Z]\d*', re.IGNORECASE)
_RE_LC_DECIMAL = re.compile(r'^[A-Z]{1,3}\s*\d{1,4}\s*\.\d+', re.IGNORECASE)
_RE_LC_CUTTER_NOSEP = re.compile(r'^[A-Z]{1,3}\d{1,4}[A-Z]\d*', re.IGNORECASE)
_RE_LC_SIMPLE = re.compile(r'^[A-Z]{1,3}\s*\d{1,4}(\s|$)', re.IGNORECASE)

# Catch-all AV and local patterns (in _classify_call_number)
_RE_AV_TAIL = re.compile(r'^(DVD|VHS|CD|VID|TAPE|VIDEO)\s*\d', re.IGNORECASE)
_RE_AV_CIRC = re.compile(r'^(Circ|Arch)\s*(CD|DVD|Video|VHS)', re.IGNORECASE)
_RE_LOCAL_2DIGIT = re.compile(r'^\d{2}\s+[A-Z]\d')
_RE_LOCAL_DATE = re.compile(r'^\d{4}-\d{1,2}-\d{1,2}')
_RE_LOCAL_ACCESSION = re.compile(r'^\d{4,}-\d+')
_RE_LOCAL_NOTATION = re.compile(r'^[*#]')


def is_not_a_call_number(cn):
    """
    Detect notes, instructions, test data, and other non-call-number data
    that has been entered in call number fields.

    Returns a category string describing what the data is:
        'public_note' — patron-facing instructions (should be in $z)
        'staff_note' — cataloging/processing notes (should be in $x)
        'equipment' — not an information resource
        'format_descriptor' — standalone format word (e.g., "DVD")
        'test_data' — placeholder/test values
        None — appears to be a real call number
    """
    cn_lower = cn.lower().strip()

    # Exact match test/placeholder values
    if cn_lower in _TEST_VALUES:
        return 'test_data'

    # Format descriptors WITHOUT numbers (just the format name alone)
    if cn_lower in _FORMAT_ONLY:
        return 'format_descriptor'

    # Equipment and supplies WITHOUT identifiers — standalone descriptions
    # like "Logitech Headset" (no number = not a call number).
    # Items WITH identifiers ("Apple Mouse #5", "TOOLKIT#1") are shelving
    # control numbers and should fall through to normal classification.
    if set(cn_lower.split()) & _EQUIPMENT_WORDS and not _RE_DIGIT.search(cn):
        return 'equipment'

    # Punctuation-only placeholders
    if _PUNCTUATION_ONLY_RE.match(cn):
        return 'test_data'

    # Pattern-based detection — check public notes first, then staff notes
    for pattern in _PUBLIC_NOTE_PATTERNS:
        if pattern.search(cn):
            return 'public_note'

    for pattern in _STAFF_NOTE_PATTERNS:
        if pattern.search(cn):
            return 'staff_note'

    return None


def is_av_shelving_number(cn):
    """
    Detect AV format shelving numbers.
    
    These are call numbers that use media format (CD, DVD, VHS, etc.) 
    as the primary organization, followed by an accession/control number.
    
    Context is important:
    - "CD ROM" alone = format descriptor, NOT a call number
    - "CD ROM 003" = shelving number (format + accession number)
    - "CD 1811" = shelving number (format + number)
    - "BRL CD ROM 071" = shelving number (collection + format + number)
    
    Note: CD and DVD are also valid LC class letters, but LC call numbers
    have different structure (class + number + cutter, e.g., "CD921 .S65")
    """
    # Pattern 1: Simple format + number (CD 1811, DVD 456, VHS-937, DVD-14)
    # Allows space or hyphen between format and number.
    # BUT NOT if followed by a cutter pattern (dot + letter), which means LC
    # e.g., "CD 3960 .P9" is LC class CD (Diplomatics), not an AV disc
    if _RE_AV_SIMPLE.match(cn):
        # Check if it looks like LC (has a cutter after the number)
        if not _RE_AV_LC_CHECK.match(cn):
            return True

    # Pattern 2: Format + ROM + number (CD ROM 003, DVD ROM 001)
    if _RE_AV_ROM.match(cn):
        return True

    # Pattern 3: Collection prefix + format (+ optional ROM) + number
    # Examples: "BRL CD ROM 071", "MUS DVD 015"
    if _RE_AV_PREFIX_FORMAT.match(cn):
        return True

    # Pattern 4: Video/Videotape + format/disc + number (with optional prefix)
    # Examples: "DSI Video CD 18", "CohenLib Video disc 110",
    #           "MusLib Video- disc MD56", "MusLib Video- recording MV74",
    #           "CohenLib Video- tape 440", "CohenLib Videotape 264"
    if _RE_AV_VIDEO.search(cn):
        return True

    # Pattern 5: VIDEO CASSETTE + number
    # Examples: "VIDEO CASSETTE 2199", "VIDEO CASSETTE 2198"
    if _RE_AV_CASSETTE.match(cn):
        return True

    # Pattern 6: Fiche/microfiche + number
    # Examples: "Fiche 414", "Microcard 5067"
    if _RE_AV_FICHE.match(cn):
        return True

    # Pattern 7: Microfilm + format code + number
    # Examples: "Microfilm MF 400"
    if _RE_AV_MICRO_FORMAT.match(cn):
        return True

    # Pattern 8: Recording + accession code (with optional prefix)
    # Examples: "MusLib Recording CD1116"
    if _RE_AV_RECORDING.search(cn):
        return True

    # Pattern 9: Music CD/format + number (without standard prefix)
    # Examples: "Music CD no.8", "CD Rhymes"
    if _RE_AV_MUSIC.match(cn):
        return True

    return False


def is_local_collection_scheme(cn):
    """
    Detect local collection schemes with recognizable patterns.
    
    These are institution-specific shelving systems that use:
    - Collection prefix + accession/control number
    - Often with hyphenated numbers
    
    Examples from CUNY:
    - BRL 200-11, BRL 201-108 (Queens College)
    - BRLV 201-08 (Queens College - "BRL Video")
    
    Returns: (is_match, confidence, note) or (False, None, None)
    """
    # Pattern: 2-5 letter prefix + hyphenated number (like BRL 200-11)
    if _RE_LOCAL_HYPHEN.match(cn):
        return True, 'Medium', 'Local collection scheme (prefix + hyphenated number)'

    # Pattern: 2-5 letter prefix + simple number (like BRLV 207)
    if _RE_LOCAL_PREFIX_NUM.match(cn):
        return True, 'Low', 'Possible local collection scheme (prefix + number)'
    
    return False, None, None


def extract_class_letters(cn):
    """Extract leading class letters from a call number."""
    match = _RE_CLASS_LETTERS.match(cn)
    return match.group(1).upper() if match else None


def is_valid_lc_class(letters):
    """Check if letters are a valid LC classification."""
    return bool(letters) and letters in LC_VALID_CLASSES


def is_sudoc(cn):
    """
    Detect SuDoc (Superintendent of Documents) classification.

    The colon (:) is the strongest single indicator of SuDoc. A call number
    containing a colon is almost always SuDoc. Very rarely, a colon may
    appear in LC call numbers — specifically in geographic cutter notation
    for subordinate locations (LC Tables G1548-G9804).

    In SuDoc, the colon appears in the class/stem area:
        A 1.10:976, Y 4.J 89/1:S 53/5, HE 20.3152:P 94
    In LC, the colon appears within a cutter for geographic subdivision:
        G3424 .A35:2C3 (Canadian Forces Base at Aldershot)
        G1778 .P4:3C8 (Curitiba, Brazil)

    The function distinguishes these by checking whether the colon follows
    an LC cutter pattern (.LetterDigits:DigitsLetterDigits).
    """
    if ':' not in cn:
        return False

    # Exclude LC geographic cutter colons — the colon is inside a cutter
    # for a subordinate location, not a SuDoc separator.
    # Pattern: LC class + number + .Letter+Digits : Digits+Letter+Digits
    if _RE_LC_CUTTER_COLON.match(cn):
        return False

    # Agency letters + number.anything + colon
    # Allow slashes, hyphens, digits, letters, and spaces between dot and colon
    if _RE_SUDOC.match(cn):
        return True

    return False


def is_dewey(cn):
    """
    Detect Dewey Decimal classification.
    
    Dewey starts with exactly 3 digits, with several format variations:
    - 394.26 (with decimal)
    - 398.2 C198T (decimal + Cutter)
    - 394 S847G (compact format, no decimal, just Cutter)
    
    Note: Three digits repeated (like "102 102") are typically local schemes.
    
    Returns: (is_match, confidence, note) or (False, None, None)
    """
    # 3 digits with decimal
    if _RE_DEWEY_DECIMAL.match(cn):
        return True, 'High', 'Dewey with decimal'

    # 3 digits + Cutter (no decimal)
    match = _RE_DEWEY_CUTTER.match(cn)
    if match:
        dewey_num = match.group(1)
        # Make sure it's not a repeated number (like "102 102")
        if not cn.startswith(f"{dewey_num} {dewey_num}"):
            return True, 'High', 'Dewey with Cutter'

    # 3 digits + author abbreviation (no decimal, no standard cutter)
    # e.g., "861 Bro 3-5" — Dewey class + truncated author name
    match = _RE_DEWEY_AUTHOR.match(cn)
    if match:
        dewey_num = match.group(1)
        if not cn.startswith(f"{dewey_num} {dewey_num}"):
            return True, 'Medium', 'Dewey with author abbreviation'

    return False, None, None


# =============================================================================
# SHELVING PREFIX STRIPPING
# =============================================================================

# Common shelving prefixes that appear before the actual classification.
# These are MARC 852 $k (call number prefix) values — they describe WHERE
# or HOW an item is shelved, not WHAT it's classified as.
# They normally belong in $k but sometimes get concatenated into $h or
# the Permanent Call Number display field.
# Sorted longest-first so "REFERENCE" matches before "REF", "PERIODICALS"
# before "PERIODICAL", etc.
SHELVING_PREFIXES = [
    'PERIODICALS', 'PERIODICAL', 'DISSERTATION', 'JUVENILE', 'REFERENCE',
    'OVERSIZE', 'RESERVE', 'SERIALS', 'SERIAL', 'THESIS', 'QUARTO',
    'FOLIO', 'SPEC', 'DOCS', 'JUV', 'PER', 'REF',
]


def strip_shelving_prefix(cn):
    """
    Strip known shelving prefixes ($k values) from the beginning of a call number.

    Requires the prefix to be followed by a space (word boundary) so that
    "REF" doesn't match inside "REFERENCE" or "RESERVED".

    Returns: (stripped_cn, prefix_found)
        - stripped_cn: the call number with the prefix removed, or the original
        - prefix_found: the prefix that was stripped, or None

    Examples:
        "OVERSIZE G 3860 1994 .H37" → ("G 3860 1994 .H37", "OVERSIZE")
        "DOCS Y 1.1/5:108-408" → ("Y 1.1/5:108-408", "DOCS")
        "E 185 .5 B58" → ("E 185 .5 B58", None)
        "REFERENCE QA76 .B3" → ("QA76 .B3", "REFERENCE")
        "Periodical QA76.73 .P98" → ("QA76.73 .P98", "PERIODICAL")
    """
    cn_upper = cn.upper()
    for prefix in SHELVING_PREFIXES:
        if cn_upper.startswith(prefix + ' '):
            rest = cn[len(prefix):].lstrip()
            if rest:
                return rest, prefix
    return cn, None


# =============================================================================
# LAC (LIBRARY AND ARCHIVES CANADA) DETECTION
# =============================================================================

def is_lac(cn):
    """
    Detect Library and Archives Canada classification.

    LAC uses two ranges that are structurally identical to LC but not
    part of the LC schedule:
    - FC: Canadian history (FC is not a valid LC class)
    - PS8000+: Canadian literature (PS is valid LC, but 8000+ is LAC)

    Returns: (is_match, confidence, note) or (False, None, None)
    """
    # FC + number = LAC Canadian history
    if _RE_LAC_FC.match(cn):
        return True, 'High', 'LAC class FC (Canadian history)'

    # PS + number >= 8000 = LAC Canadian literature
    ps_match = _RE_LAC_PS.match(cn)
    if ps_match:
        ps_num = int(ps_match.group(1))
        if ps_num >= 8000:
            return True, 'High', 'LAC class PS8000+ (Canadian literature)'

    return False, None, None


# =============================================================================
# CLASSIFICATION HELPER
# =============================================================================

def _classify_call_number(cn_stripped):
    """
    Classify a call number string (already stripped of any $k prefix).

    Returns (indicator, scheme, confidence, note) if a classification is
    found, or None if the pattern is not recognized.
    """
    # === SUDOC ===
    if is_sudoc(cn_stripped):
        return '3', 'Superintendent of Documents', 'High', 'SuDoc pattern (colon separator)'

    # Y class is always SuDoc (Congressional), never LC
    if _RE_SUDOC_Y.match(cn_stripped):
        return '3', 'Superintendent of Documents', 'High', 'SuDoc Y class (Congressional)'

    # === NLM ===
    class_letters = extract_class_letters(cn_stripped)
    if class_letters and _RE_NLM_CLASS.match(class_letters):
        if _RE_NLM_NUM.match(cn_stripped):
            return '2', 'National Library of Medicine', 'High', 'NLM class (QS-QZ or W)'

    # === LAC (Library and Archives Canada) ===
    is_lac_match, conf, note = is_lac(cn_stripped)
    if is_lac_match:
        return '7', 'Library and Archives Canada', conf, note

    # === DEWEY ===
    is_dew, conf, note = is_dewey(cn_stripped)
    if is_dew:
        return '1', 'Dewey Decimal', conf, note

    # === LOCAL RESERVE LABELS ===
    # Abbreviated title + year + edition (e.g., "Am 2014 4th Ed",
    # "CJ 2017 3rd Ed"). These are local shelving labels for course
    # reserves, not classification numbers. Check before LC to prevent
    # false matches on valid LC class letters.
    if _RE_RESERVE_EDITION.match(cn_stripped):
        return '8', 'Other scheme', 'Medium', 'Local shelving label (title abbreviation + edition)'
    # Letters + small number + year, no cutter (e.g., "RM 30 2016").
    # Real LC call numbers with a date almost always have a cutter
    # between the class number and the date.
    if _RE_RESERVE_YEAR.match(cn_stripped):
        return '8', 'Other scheme', 'Medium', 'Local shelving label (title abbreviation + year)'

    # === LC CLASSIFICATION ===
    if class_letters and is_valid_lc_class(class_letters):
        # CIP/preliminary LC (MLCS pattern)
        if _RE_LC_MLCS.search(cn_stripped):
            return '0', 'Library of Congress', 'Low', 'LC (CIP/preliminary — MLCS number)'
        # LC with attached cutter (no space between class number and cutter)
        if _RE_LC_CUTTER_ATTACHED.match(cn_stripped):
            return '0', 'Library of Congress', 'High', 'LC with cutter'
        # LC with cutter (space between class number and cutter)
        if _RE_LC_CUTTER_SPACE.match(cn_stripped):
            return '0', 'Library of Congress', 'High', 'LC with cutter'
        # LC with number + date + cutter
        if _RE_LC_DATE_CUTTER.match(cn_stripped):
            return '0', 'Library of Congress', 'High', 'LC with date and cutter'
        # LC with decimal but no cutter
        if _RE_LC_DECIMAL.match(cn_stripped):
            return '0', 'Library of Congress', 'Medium', 'LC class with decimal'
        # LC with cutter directly attached (no dot, no space)
        # e.g., PQ2402A3, M1510S3.8, PR3716F55L, N6953R4
        if _RE_LC_CUTTER_NOSEP.match(cn_stripped):
            return '0', 'Library of Congress', 'Medium', 'LC with cutter (no separator)'
        # Simple LC (just class and number)
        if _RE_LC_SIMPLE.match(cn_stripped):
            return '0', 'Library of Congress', 'Medium', 'LC class and number'

    # === LOCAL COLLECTION SCHEMES ===
    is_local, conf, note = is_local_collection_scheme(cn_stripped)
    if is_local:
        return '4', 'Shelving control number', conf, note

    # === OTHER AV PATTERNS ===
    if _RE_AV_TAIL.match(cn_stripped):
        return '4', 'Shelving control number', 'High', 'AV format shelving'
    if _RE_AV_CIRC.match(cn_stripped):
        return '4', 'Shelving control number', 'High', 'AV circulation shelving'

    # === LOCAL SCHEMES ===
    if _RE_LOCAL_2DIGIT.match(cn_stripped):
        return '4', 'Shelving control number', 'Medium', 'Local shelving (2-digit prefix)'
    if _RE_LOCAL_DATE.match(cn_stripped):
        return '4', 'Shelving control number', 'High', 'Date-based shelving'
    if _RE_LOCAL_ACCESSION.match(cn_stripped):
        return '4', 'Shelving control number', 'Medium', 'Accession number'

    # === LOCAL NOTATION ===
    if _RE_LOCAL_NOTATION.match(cn_stripped):
        return '4', 'Shelving control number', 'Low', 'Local notation'

    return None


# =============================================================================
# MAIN CLASSIFICATION FUNCTION
# =============================================================================

def categorize_call_number(call_num, from_j=False, j_combined=False,
                           j_conflict=False, institution=None):
    """
    Analyze a call number and suggest the appropriate 852 first indicator.

    Always classifies by content, regardless of which subfield the data
    came from. If the content is in $j but looks like a standard
    classification (LC, Dewey, etc.), the script flags it as a subfield
    error rather than accepting it as shelving control.

    Args:
        call_num: The call number string to classify.
        from_j: True if the call number came from MARC 852 $j. If the
                content is not a shelving control number, $j was used
                by mistake and a subfield change is noted.
        j_combined: True if $j was merged with $h/$i (miscoded cutter).
        j_conflict: True if $h has a classification AND $j has a separate
                    shelving control number (two schemes in one 852 field).
        institution: Institution name for per-campus analysis.

    Returns: (indicator, classification_type, confidence, note, subfield_changes)

    Indicator values:
        0 = Library of Congress
        1 = Dewey Decimal
        2 = National Library of Medicine
        3 = Superintendent of Documents
        4 = Shelving control number
        5 = Title
        7 = Source specified in $2
        8 = Other scheme
        N/A = Not a call number (data quality issue)
        blank = Unknown/missing
    """
    if pd.isna(call_num) or not call_num:
        return 'blank', 'Unknown', 'Low', 'Missing call number', ''

    cn = str(call_num).strip()
    subfield_changes = []

    # === NON-CALL-NUMBERS ===
    # Check this BEFORE the $j override — "DVD" alone is a format
    # descriptor even if a trusted institution put it in $j.
    not_cn = is_not_a_call_number(cn)
    if not_cn:
        note_map = {
            'public_note': 'Public note',
            'staff_note': 'Staff/cataloging note',
            'equipment': 'Equipment description — may be intentional for circulation',
            'format_descriptor': 'Format descriptor only',
            'test_data': 'Test/placeholder data',
        }
        sf_map = {
            'public_note': 'Move to $z',
            'staff_note': 'Move to $x',
        }
        sf = sf_map.get(not_cn, '')
        return 'N/A', 'Not a call number', 'High', note_map.get(not_cn, 'Not a call number'), sf

    # === AV SHELVING NUMBERS ===
    if is_av_shelving_number(cn):
        return '4', 'Shelving control number', 'High', 'AV format shelving', ''

    # === CHECK FOR PREFIX-ONLY CALL NUMBERS ===
    # Words like "Periodical", "Thesis", "Reference" are $k (call number
    # prefix) values. When they appear alone with no classification after
    # them, the indicator is ambiguous.
    if cn.upper().strip() in SHELVING_PREFIXES:
        prefix_word = cn.strip()
        return ('8', 'Other scheme', 'Low',
                "Prefix only — no classification follows. Could be shelving control (4) or other scheme (8)",
                f"Move '{prefix_word}' to $k")

    # === STRIP SHELVING PREFIXES ===
    # Try stripping $k prefixes (OVERSIZE, DOCS, PERIODICAL, THESIS, etc.)
    # for classification. If a prefix is found, classify the remainder.
    cn_stripped, prefix_found = strip_shelving_prefix(cn)

    # Classify the (possibly stripped) call number
    result = _classify_call_number(cn_stripped)

    if result:
        indicator, scheme, conf, note = result
        if prefix_found:
            subfield_changes.append(f"Move '{prefix_found}' to $k")
    elif prefix_found:
        # === PREFIX STRIPPED BUT REMAINDER NOT CLASSIFIED ===
        indicator, scheme, conf, note = ('8', 'Other scheme', 'Low',
            f"Remainder '{cn_stripped}' not a standard classification. "
            f"Could be shelving control (4) or other scheme (8)")
        subfield_changes.append(f"Move '{prefix_found}' to $k")
    else:
        # === UNRECOGNIZED ===
        indicator, scheme, conf, note = '8', 'Other scheme', 'Low', 'Pattern not recognized - review recommended'

    # If the value came from $j, check whether $j is appropriate.
    # $j is for shelving control numbers (indicator 4). If the content
    # is a standard classification, $j was used by mistake.
    if from_j and indicator != '4':
        subfield_changes.append('Move $j to $h')

    # If $j was combined with $h/$i, the cutter was miscoded in $j
    # instead of $i. Note the subfield error.
    if j_combined:
        subfield_changes.append('Move $j cutter to $i')

    # If $h has a classification and $j has a separate shelving control
    # number, flag for review — two schemes in one 852 field.
    if j_conflict:
        indicator = 'N/A'
        note = f'{scheme} in $h but shelving control number in $j — review needed'
        conf = 'Review'

    return indicator, scheme, conf, note, '; '.join(subfield_changes)


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

def create_excel_output(df, output_path):
    """Create formatted Excel workbook with analysis results."""
    
    wb = Workbook()

    # Styles
    data_font = Font(name='Arial', size=12)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    conf_colors = {
        'High': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'Medium': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'Low': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    }
    not_cn_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    
    # === SHEET 1: Main data ===
    ws_data = wb.active
    ws_data.title = "852 Field Analysis"

    # Force ID columns to string to prevent scientific notation
    for col in ['MMS Id', 'Holdings ID']:
        if col in df.columns:
            df[col] = df[col].astype(str)

    change_yes_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    change_no_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    change_review_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

    # ID columns (MMS Id = 8, Holdings ID = 9) — stored as text
    id_col_indices = {8, 9}

    headers = [
        'Permanent Call Number', 'Extracted Call Number', 'Permanent Call Number Type',
        '852 MARC', 'Normalized Call Number', 'Institution Name', 'Library Name',
        'MMS Id', 'Holdings ID', 'Suppressed',
        'Current Indicator', 'Suggested Indicator', 'Change Needed',
        'Classification Type', 'Confidence', 'Subfield Changes', 'Notes'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    for excel_row, (_, row) in enumerate(df.iterrows(), 2):
        data = [
            row['Permanent Call Number'], row['Extracted Call Number'],
            row['Permanent Call Number Type'], row['852 MARC'],
            row['Normalized Call Number'], row['Institution Name'], row['Library Name'],
            row['MMS Id'], row['Holdings ID'], row['Suppressed'],
            row['Current Indicator'], row['Suggested Indicator'], row['Change Needed'],
            row['Classification Type'], row['Confidence'],
            row['Subfield Changes'], row['Notes']
        ]
        for col_idx, value in enumerate(data, 1):
            cell = ws_data.cell(row=excel_row, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            if col_idx in id_col_indices:
                cell.number_format = '@'
            # Change Needed column coloring (column 13)
            if col_idx == 13:
                if value == 'Yes':
                    cell.fill = change_yes_fill
                elif value == 'No':
                    cell.fill = change_no_fill
                elif value == 'Review':
                    cell.fill = change_review_fill
            # Confidence column coloring (column 15)
            if col_idx == 15 and value in conf_colors:
                cell.fill = conf_colors[value]
            # Not a call number highlighting (column 14)
            if col_idx == 14 and value == 'Not a call number':
                cell.fill = not_cn_fill
                cell.font = Font(name='Arial', size=12, bold=True)

    col_widths = {
        'A': 35, 'B': 30, 'C': 25, 'D': 60, 'E': 45,
        'F': 30, 'G': 30, 'H': 20, 'I': 20, 'J': 15,
        'K': 18, 'L': 18, 'M': 15,
        'N': 28, 'O': 12, 'P': 40, 'Q': 55
    }
    for col, width in col_widths.items():
        ws_data.column_dimensions[col].width = width

    ws_data.freeze_panes = 'A2'
    ws_data.auto_filter.ref = f"A1:Q{len(df) + 1}"
    
    # === SHEET 2: Statistics ===
    ws_stats = wb.create_sheet("Statistics")
    
    summary_indicator = df.groupby(['Suggested Indicator', 'Classification Type']).size().reset_index(name='Count')
    summary_indicator['Percentage'] = (summary_indicator['Count'] / len(df) * 100).round(2)
    summary_indicator = summary_indicator.sort_values('Count', ascending=False)
    
    summary_confidence = df.groupby('Confidence').size().reset_index(name='Count')
    summary_confidence['Percentage'] = (summary_confidence['Count'] / len(df) * 100).round(2)
    
    summary_institution = df.groupby('Institution Name').size().reset_index(name='Count')
    summary_institution = summary_institution.sort_values('Count', ascending=False)
    
    row = 1
    ws_stats.cell(row=row, column=1, value="852 First Indicator Analysis - Statistics")
    ws_stats.cell(row=row, column=1).font = Font(name='Arial', size=14, bold=True)
    row += 2
    
    ws_stats.cell(row=row, column=1, value="Overall Summary")
    ws_stats.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True)
    row += 1
    ws_stats.cell(row=row, column=1, value="Total Records:").font = data_font
    ws_stats.cell(row=row, column=2, value=len(df)).font = data_font
    row += 1
    ws_stats.cell(row=row, column=1, value="Records with Extracted Call Number:").font = data_font
    ws_stats.cell(row=row, column=2, value=df['Extracted Call Number'].notna().sum()).font = data_font
    row += 2
    
    ws_stats.cell(row=row, column=1, value="By Classification Type")
    ws_stats.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True)
    row += 1
    
    for col_idx, header in enumerate(['Suggested Indicator', 'Classification Type', 'Count', 'Percentage'], 1):
        cell = ws_stats.cell(row=row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    row += 1
    
    for _, data_row in summary_indicator.iterrows():
        for col_idx, val in enumerate([data_row['Suggested Indicator'],
                                       data_row['Classification Type'],
                                       data_row['Count']], 1):
            c = ws_stats.cell(row=row, column=col_idx, value=val)
            c.font = data_font
            c.border = thin_border
        cell = ws_stats.cell(row=row, column=4, value=data_row['Percentage'] / 100)
        cell.number_format = '0.00%'
        cell.font = data_font
        cell.border = thin_border
        row += 1
    row += 1
    
    ws_stats.cell(row=row, column=1, value="By Confidence Level")
    ws_stats.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True)
    row += 1
    
    for col_idx, header in enumerate(['Confidence', 'Count', 'Percentage'], 1):
        cell = ws_stats.cell(row=row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    row += 1
    
    for _, data_row in summary_confidence.iterrows():
        cell1 = ws_stats.cell(row=row, column=1, value=data_row['Confidence'])
        cell1.font = data_font
        cell1.border = thin_border
        if data_row['Confidence'] in conf_colors:
            cell1.fill = conf_colors[data_row['Confidence']]
        c2 = ws_stats.cell(row=row, column=2, value=data_row['Count'])
        c2.font = data_font
        c2.border = thin_border
        cell = ws_stats.cell(row=row, column=3, value=data_row['Percentage'] / 100)
        cell.number_format = '0.00%'
        cell.font = data_font
        cell.border = thin_border
        row += 1
    row += 1
    
    ws_stats.cell(row=row, column=1, value="By Institution")
    ws_stats.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True)
    row += 1
    
    for col_idx, header in enumerate(['Institution', 'Count'], 1):
        cell = ws_stats.cell(row=row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    row += 1
    
    for _, data_row in summary_institution.iterrows():
        c1 = ws_stats.cell(row=row, column=1, value=data_row['Institution Name'])
        c1.font = data_font
        c1.border = thin_border
        c2 = ws_stats.cell(row=row, column=2, value=data_row['Count'])
        c2.font = data_font
        c2.border = thin_border
        row += 1
    
    ws_stats.column_dimensions['A'].width = 50
    ws_stats.column_dimensions['B'].width = 30
    ws_stats.column_dimensions['C'].width = 15
    ws_stats.column_dimensions['D'].width = 12
    
    # === SHEET 3: By Institution ===
    ws_inst = wb.create_sheet("By Institution")
    
    crosstab = pd.crosstab(df['Institution Name'], df['Suggested Indicator'], margins=True, dropna=False)
    
    ws_inst.cell(row=1, column=1, value="Classification by Institution")
    ws_inst.cell(row=1, column=1).font = Font(name='Arial', size=14, bold=True)
    
    row = 3
    ws_inst.cell(row=row, column=1, value="Count by Institution and Suggested Indicator")
    ws_inst.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True)
    row += 1
    
    headers = ['Institution'] + [str(c) for c in crosstab.columns]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_inst.cell(row=row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    row += 1
    
    for inst in crosstab.index:
        c = ws_inst.cell(row=row, column=1, value=inst)
        c.font = data_font
        c.border = thin_border
        for col_idx, ind in enumerate(crosstab.columns, 2):
            c = ws_inst.cell(row=row, column=col_idx, value=int(crosstab.loc[inst, ind]))
            c.font = data_font
            c.border = thin_border
        row += 1
    
    row += 2
    ws_inst.cell(row=row, column=1, value="Sample 'Other Scheme' and 'Unknown' Entries by Institution")
    ws_inst.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True)
    row += 1
    ws_inst.cell(row=row, column=1, value="(These may be local schemes that vary by campus)")
    ws_inst.cell(row=row, column=1).font = Font(name='Arial', size=12, italic=True)
    row += 2
    
    other_unknown = df[df['Suggested Indicator'].isin(['8', 'blank', 'N/A'])]
    for inst in other_unknown['Institution Name'].value_counts().head(10).index:
        ws_inst.cell(row=row, column=1, value=inst).font = Font(name='Arial', size=12, bold=True)
        row += 1
        samples = other_unknown[other_unknown['Institution Name'] == inst]['Extracted Call Number'].dropna().unique()[:8]
        for sample in samples:
            ws_inst.cell(row=row, column=2, value=str(sample)[:60]).font = data_font
            row += 1
        row += 1
    
    ws_inst.column_dimensions['A'].width = 50
    ws_inst.column_dimensions['B'].width = 40
    for col in 'CDEFGHIJKLMNO':
        ws_inst.column_dimensions[col].width = 10
    
    wb.save(output_path)


# =============================================================================
# HTML REPORT
# =============================================================================

# HTML template with placeholder tokens (__TOKEN__) that get replaced by Python.
# Using placeholders instead of f-strings avoids escaping all the {} in CSS/JS.

_HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>852 Field Analysis Report</title>
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }

body {
    font-family: Arial, sans-serif;
    font-size: 12pt;
    background: #f5f5f5;
    color: #333;
    line-height: 1.5;
}

header {
    background: #4472C4;
    color: white;
    padding: 24px 32px;
}
header h1 { font-size: 18pt; }
header .subtitle { font-size: 11pt; opacity: 0.9; margin-top: 4px; }

/* Dashboard cards */
#dashboard {
    display: flex;
    gap: 16px;
    padding: 24px 32px;
    flex-wrap: wrap;
}
.card {
    flex: 1;
    min-width: 180px;
    padding: 16px 20px;
    border-radius: 8px;
    text-align: center;
}
.card .big-number { font-size: 28pt; font-weight: bold; }
.card .label { font-size: 10pt; margin-top: 4px; }
.card-yes    { background: #FFC7CE; color: #9C0006; }
.card-review { background: #FFEB9C; color: #9C6500; }
.card-no     { background: #C6EFCE; color: #006100; }
.card-info   { background: #D6E4F0; color: #2C5282; }

/* Filter controls */
#controls {
    padding: 16px 32px;
    background: white;
    border-bottom: 1px solid #ddd;
    display: flex;
    gap: 12px;
    flex-wrap: wrap;
    align-items: flex-end;
}
.filter-group { display: flex; flex-direction: column; gap: 2px; }
.filter-group label {
    font-size: 9pt;
    font-weight: bold;
    color: #555;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}
#controls select, #controls input {
    font-family: Arial, sans-serif;
    font-size: 11pt;
    padding: 6px 10px;
    border: 1px solid #ccc;
    border-radius: 4px;
    height: 36px;
    box-sizing: border-box;
}
#search-box { min-width: 200px; }

/* Pagination */
#pagination {
    display: flex;
    align-items: center;
    gap: 12px;
    margin-top: 12px;
}
#pagination button {
    font-family: Arial, sans-serif;
    font-size: 11pt;
    padding: 6px 16px;
    border: 1px solid #ccc;
    border-radius: 4px;
    background: white;
    cursor: pointer;
}
#pagination button:hover:not(:disabled) { background: #e8e8e8; }
#pagination button:disabled { opacity: 0.4; cursor: default; }
#page-info { font-size: 10pt; color: #555; }

/* Table */
#table-container {
    padding: 0 32px 32px;
    overflow-x: auto;
}
table {
    width: 100%;
    border-collapse: collapse;
    margin-top: 16px;
    background: white;
}
th {
    background: #4472C4;
    color: white;
    padding: 10px 12px;
    text-align: left;
    cursor: pointer;
    user-select: none;
    white-space: nowrap;
    font-size: 11pt;
    position: sticky;
    top: 0;
    z-index: 10;
    border-bottom: 3px solid #4472C4;
    transition: background 0.15s;
}
th:hover { background: #365FA0; }
th .sort-arrow {
    margin-left: 4px;
    font-size: 9pt;
    opacity: 0.4;
}
th.sorted { background: #365FA0; border-bottom: 3px solid #FFC000; }
th.sorted .sort-arrow { opacity: 1; }
td {
    padding: 8px 12px;
    border-bottom: 1px solid #e5e5e5;
    font-size: 11pt;
}
tr:hover { background: #f0f4ff; }

/* Color-coded cells */
td.change-yes    { background: #FFC7CE; color: #9C0006; font-weight: bold; }
td.change-no     { background: #C6EFCE; color: #006100; }
td.change-review { background: #FFEB9C; color: #9C6500; font-weight: bold; }
td.conf-high   { background: #C6EFCE; color: #006100; }
td.conf-medium { background: #FFEB9C; color: #9C6500; }
td.conf-low    { background: #FFC7CE; color: #9C0006; }
td.not-cn { color: #CC0000; font-weight: bold; }

/* Primo links */
a { color: #2B5797; }
a:visited { color: #6B4C9A; }
a:hover { color: #1A3A6B; text-decoration: underline; }

/* Row count and footer */
#row-count {
    margin-top: 12px;
    font-size: 10pt;
    color: #666;
}
footer {
    padding: 24px 32px;
    text-align: center;
    color: #888;
    font-size: 10pt;
}
#export-csv {
    font-family: Arial, sans-serif;
    font-size: 11pt;
    padding: 10px 24px;
    background: #4472C4;
    color: white;
    border: none;
    border-radius: 4px;
    cursor: pointer;
    margin-bottom: 12px;
}
#export-csv:hover { background: #365FA0; }
#reset-filters {
    font-family: Arial, sans-serif;
    font-size: 10pt;
    padding: 6px 16px;
    background: #4472C4;
    color: white;
    border: none;
    border-radius: 4px;
    cursor: pointer;
    align-self: flex-end;
}
#reset-filters:hover { background: #365FA0; }
</style>
</head>
<body>

<header>
    <h1>852 First Indicator Analysis</h1>
    <p class="subtitle">Generated __DATE__ &mdash; __TOTAL__ records analyzed, __COUNT_CHANGES__ shown below</p>
</header>

<section id="dashboard">
    <div class="card card-yes">
        <div class="big-number" id="dash-yes">__COUNT_YES__</div>
        <div class="label">changes needed</div>
    </div>
    <div class="card card-review">
        <div class="big-number" id="dash-review">__COUNT_REVIEW__</div>
        <div class="label">need manual review</div>
    </div>
    <div class="card card-no">
        <div class="big-number" id="dash-no">__COUNT_NO__</div>
        <div class="label">already correct</div>
    </div>
    <div class="card card-info">
        <div class="big-number" id="dash-libraries">__COUNT_LIBRARIES__</div>
        <div class="label">libraries</div>
    </div>
</section>

<section id="controls">
    <div class="filter-group">
        <label for="filter-library">Library</label>
        <select id="filter-library"><option value="all">All Libraries</option></select>
    </div>
    <div class="filter-group">
        <label for="filter-change">Change Needed</label>
        <select id="filter-change">
            <option value="all">All</option>
            <option value="yes">Yes</option>
            <option value="review">Review</option>
        </select>
    </div>
    <div class="filter-group">
        <label for="filter-confidence">Confidence</label>
        <select id="filter-confidence">
            <option value="all">All</option>
            <option value="High">High</option>
            <option value="Medium">Medium</option>
            <option value="Low">Low</option>
        </select>
    </div>
    <div class="filter-group">
        <label for="filter-class-type">Classification</label>
        <select id="filter-class-type"><option value="all">All</option></select>
    </div>
    <div class="filter-group">
        <label for="filter-subfield">Subfield Changes</label>
        <select id="filter-subfield"><option value="all">All</option></select>
    </div>
    <div class="filter-group">
        <label for="filter-suppressed">Suppressed</label>
        <select id="filter-suppressed">
            <option value="all">All</option>
            <option value="no">Not suppressed</option>
            <option value="yes">Suppressed</option>
        </select>
    </div>
    <div class="filter-group">
        <label for="search-box">Search</label>
        <input type="text" id="search-box" placeholder="Search call numbers, notes...">
    </div>
    <div class="filter-group">
        <button id="reset-filters" onclick="resetFilters()">Reset Filters</button>
    </div>
</section>

<section id="table-container">
    <table id="data-table">
        <thead>
            <tr>
                <th onclick="sortTable(0)">Library <span class="sort-arrow"></span></th>
                <th onclick="sortTable(1)">Call Number <span class="sort-arrow"></span></th>
                <th onclick="sortTable(2)">Current <span class="sort-arrow"></span></th>
                <th onclick="sortTable(3)">Suggested <span class="sort-arrow"></span></th>
                <th onclick="sortTable(4)">Change Needed <span class="sort-arrow"></span></th>
                <th onclick="sortTable(5)">Classification <span class="sort-arrow"></span></th>
                <th onclick="sortTable(6)">Confidence <span class="sort-arrow"></span></th>
                <th onclick="sortTable(7)">Subfield Changes <span class="sort-arrow"></span></th>
                <th onclick="sortTable(8)">Notes <span class="sort-arrow"></span></th>
                <th onclick="sortTable(9)">MMS Id <span class="sort-arrow"></span></th>
                <th onclick="sortTable(10)">Holdings ID <span class="sort-arrow"></span></th>
                <th onclick="sortTable(11)">Suppressed <span class="sort-arrow"></span></th>
            </tr>
        </thead>
        <tbody id="table-body"></tbody>
    </table>
    <p id="row-count"></p>
    <div id="pagination">
        <button id="page-prev" onclick="changePage(-1)">&laquo; Previous</button>
        <span id="page-info"></span>
        <button id="page-next" onclick="changePage(1)">Next &raquo;</button>
    </div>
</section>

<footer>
    <button id="export-csv" onclick="exportCSV()">Export Filtered Data as CSV</button>
    <p>Generated by 852 First Indicator Analysis Script</p>
</footer>

<script>
// ============================================================
// DATA (injected by Python)
// ============================================================
var ALL_DATA = __JSON_DATA__;

// Column keys in display order (must match <th> order above)
var COLS = [
    'library', 'callNumber', 'currentInd', 'suggestedInd',
    'changeNeeded', 'classType', 'confidence', 'subfieldChanges',
    'notes', 'mmsId', 'holdingsId', 'suppressed'
];

// Pagination state
var PAGE_SIZE = 500;
var currentPage = 1;
var filteredData = [];

// Pre-build search text for each record (once, at load time)
for (var i = 0; i < ALL_DATA.length; i++) {
    var r = ALL_DATA[i];
    r._search = (r.callNumber + ' ' + r.classType + ' ' + r.notes + ' ' +
                  r.subfieldChanges + ' ' + r.mmsId).toLowerCase();
}

// ============================================================
// HTML HELPERS
// ============================================================
function esc(s) {
    if (!s) return '';
    return s.replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;')
            .replace(/"/g, '&quot;');
}
function escAttr(s) { return esc(s); }

function mmsIdCell(mmsId, primoCode, scopeCode) {
    if (!mmsId) return '';
    if (!primoCode) return esc(mmsId);
    var code = primoCode.toLowerCase();
    var CODE = primoCode.toUpperCase();
    var url;
    if (scopeCode) {
        // Permalink format (preferred)
        url = 'https://cuny-' + code + '.primo.exlibrisgroup.com/permalink/01CUNY_'
            + CODE + '/' + scopeCode + '/alma' + encodeURIComponent(mmsId);
    } else {
        // Fallback: raw MARC view (for schools without scope codes yet)
        url = 'https://cuny-' + code + '.primo.exlibrisgroup.com/discovery/sourceRecord'
            + '?vid=01CUNY_' + CODE + ':CUNY_' + CODE
            + '&docId=alma' + encodeURIComponent(mmsId);
    }
    return '<a href="' + escAttr(url) + '" target="_blank" title="View in Primo">'
         + esc(mmsId) + '</a>';
}

// ============================================================
// RENDER CURRENT PAGE
// ============================================================
function renderPage() {
    var tbody = document.getElementById('table-body');
    var totalPages = Math.ceil(filteredData.length / PAGE_SIZE) || 1;
    if (currentPage > totalPages) currentPage = totalPages;
    var start = (currentPage - 1) * PAGE_SIZE;
    var end = Math.min(start + PAGE_SIZE, filteredData.length);

    var html = '';
    for (var i = start; i < end; i++) {
        var r = filteredData[i];
        var changeCls = '';
        if (r.changeNeeded === 'Yes') changeCls = 'change-yes';
        else if (r.changeNeeded === 'Review') changeCls = 'change-review';
        else if (r.changeNeeded === 'No') changeCls = 'change-no';

        var confCls = '';
        if (r.confidence === 'High') confCls = 'conf-high';
        else if (r.confidence === 'Medium') confCls = 'conf-medium';
        else if (r.confidence === 'Low') confCls = 'conf-low';

        var classCls = '';
        if (r.classType === 'Not a call number') classCls = 'not-cn';

        html += '<tr>'
            + '<td>' + esc(r.library) + '</td>'
            + '<td>' + esc(r.callNumber) + '</td>'
            + '<td>' + esc(r.currentInd) + '</td>'
            + '<td>' + esc(r.suggestedInd) + '</td>'
            + '<td class="' + changeCls + '">' + esc(r.changeNeeded) + '</td>'
            + '<td class="' + classCls + '">' + esc(r.classType) + '</td>'
            + '<td class="' + confCls + '">' + esc(r.confidence) + '</td>'
            + '<td>' + esc(r.subfieldChanges) + '</td>'
            + '<td>' + esc(r.notes) + '</td>'
            + '<td>' + mmsIdCell(r.mmsId, r.primoCode, r.scopeCode) + '</td>'
            + '<td>' + esc(r.holdingsId) + '</td>'
            + '<td>' + esc(r.suppressed) + '</td>'
            + '</tr>';
    }
    tbody.innerHTML = html;

    // Update row count and pagination controls
    document.getElementById('row-count').textContent =
        'Showing ' + (filteredData.length ? start + 1 : 0) + '–' + end +
        ' of ' + filteredData.length.toLocaleString() + ' filtered records' +
        ' (' + ALL_DATA.length.toLocaleString() + ' total)';

    document.getElementById('page-info').textContent =
        'Page ' + currentPage + ' of ' + totalPages;
    document.getElementById('page-prev').disabled = (currentPage <= 1);
    document.getElementById('page-next').disabled = (currentPage >= totalPages);
}

// ============================================================
// PAGINATION
// ============================================================
function changePage(delta) {
    var totalPages = Math.ceil(filteredData.length / PAGE_SIZE) || 1;
    var newPage = currentPage + delta;
    if (newPage >= 1 && newPage <= totalPages) {
        currentPage = newPage;
        renderPage();
        // Scroll to top of table
        document.getElementById('table-container').scrollIntoView({ behavior: 'smooth' });
    }
}

// ============================================================
// FILTERING
// ============================================================
function resetFilters() {
    document.getElementById('filter-library').value = 'all';
    document.getElementById('filter-change').value = 'all';
    document.getElementById('filter-confidence').value = 'all';
    document.getElementById('filter-class-type').value = 'all';
    document.getElementById('filter-subfield').value = 'all';
    document.getElementById('filter-suppressed').value = 'all';
    document.getElementById('search-box').value = '';
    applyFilters();
}

function applyFilters() {
    var library = document.getElementById('filter-library').value;
    var changeVal = document.getElementById('filter-change').value;
    var confidence = document.getElementById('filter-confidence').value;
    var classType = document.getElementById('filter-class-type').value;
    var subfield = document.getElementById('filter-subfield').value;
    var suppressed = document.getElementById('filter-suppressed').value;
    var searchText = document.getElementById('search-box').value.toLowerCase();

    filteredData = [];
    var countYes = 0, countReview = 0, countNo = 0;
    var libSeen = {};

    for (var i = 0; i < ALL_DATA.length; i++) {
        var r = ALL_DATA[i];
        var show = true;

        if (library !== 'all' && r.library !== library) show = false;
        if (show && changeVal !== 'all') {
            if (changeVal === 'yes' && r.changeNeeded !== 'Yes') show = false;
            else if (changeVal === 'review' && r.changeNeeded !== 'Review') show = false;
        }
        if (show && confidence !== 'all' && r.confidence !== confidence) show = false;
        if (show && classType !== 'all' && r.classType !== classType) show = false;
        if (show && subfield !== 'all') {
            var sf = r.subfieldChanges || '';
            if (subfield === 'none' && sf) show = false;
            else if (subfield === 'any' && !sf) show = false;
            else if (subfield !== 'none' && subfield !== 'any' && sf.indexOf(subfield) === -1) show = false;
        }
        if (show && suppressed !== 'all') {
            var sup = (r.suppressed || '').toLowerCase();
            var isSuppressed = (sup === 'yes' || sup === 'true');
            if (suppressed === 'yes' && !isSuppressed) show = false;
            else if (suppressed === 'no' && isSuppressed) show = false;
        }
        if (show && searchText && r._search.indexOf(searchText) === -1) show = false;

        if (show) {
            filteredData.push(r);
            if (r.changeNeeded === 'Yes') countYes++;
            else if (r.changeNeeded === 'Review') countReview++;
            else if (r.changeNeeded === 'No') countNo++;
            if (r.library) libSeen[r.library] = true;
        }
    }

    // Update dashboard cards
    document.getElementById('dash-yes').textContent = countYes.toLocaleString();
    document.getElementById('dash-review').textContent = countReview.toLocaleString();
    document.getElementById('dash-no').textContent = countNo.toLocaleString();
    document.getElementById('dash-libraries').textContent = Object.keys(libSeen).length;

    // Re-apply current sort if one is active
    if (sortCol !== null) {
        var key = COLS[sortCol];
        filteredData.sort(function(a, b) {
            var va = (a[key] || '').toLowerCase();
            var vb = (b[key] || '').toLowerCase();
            if (va < vb) return sortAsc ? -1 : 1;
            if (va > vb) return sortAsc ? 1 : -1;
            return 0;
        });
    }

    // Reset to page 1 and render
    currentPage = 1;
    renderPage();
}

// ============================================================
// SORTING
// ============================================================
var sortCol = -1;
var sortAsc = true;

function sortTable(colIndex) {
    if (sortCol === colIndex) {
        sortAsc = !sortAsc;
    } else {
        sortCol = colIndex;
        sortAsc = true;
    }

    var ths = document.querySelectorAll('#data-table th');
    for (var i = 0; i < ths.length; i++) {
        var arrow = ths[i].querySelector('.sort-arrow');
        if (i === colIndex) {
            ths[i].classList.add('sorted');
            if (arrow) arrow.textContent = sortAsc ? '\\u25B2' : '\\u25BC';
        } else {
            ths[i].classList.remove('sorted');
            if (arrow) arrow.textContent = '\\u25E5';
        }
    }

    var key = COLS[colIndex];
    filteredData.sort(function(a, b) {
        var va = (a[key] || '').toLowerCase();
        var vb = (b[key] || '').toLowerCase();
        if (va < vb) return sortAsc ? -1 : 1;
        if (va > vb) return sortAsc ? 1 : -1;
        return 0;
    });

    currentPage = 1;
    renderPage();
}

// ============================================================
// CSV EXPORT (all filtered data, not just current page)
// ============================================================
function exportCSV() {
    var headers = [
        'Library Name', 'Extracted Call Number', 'Current Indicator',
        'Suggested Indicator', 'Change Needed', 'Classification Type',
        'Confidence', 'Subfield Changes', 'Notes', 'MMS Id', 'Holdings ID',
        'Suppressed'
    ];

    var csvRows = [headers.join(',')];
    for (var i = 0; i < filteredData.length; i++) {
        var r = filteredData[i];
        var vals = [
            r.library, r.callNumber, r.currentInd, r.suggestedInd,
            r.changeNeeded, r.classType, r.confidence, r.subfieldChanges,
            r.notes, r.mmsId, r.holdingsId, r.suppressed
        ];
        csvRows.push(vals.map(function(v) {
            return '"' + (v || '').replace(/"/g, '""') + '"';
        }).join(','));
    }

    var blob = new Blob([csvRows.join('\\n')], { type: 'text/csv;charset=utf-8;' });
    var url = URL.createObjectURL(blob);
    var link = document.createElement('a');
    link.href = url;
    link.download = 'call_number_changes.csv';
    link.click();
    URL.revokeObjectURL(url);
}

// ============================================================
// POPULATE DROPDOWNS
// ============================================================
function populateDropdown(selectId, values) {
    var select = document.getElementById(selectId);
    var unique = [];
    var seen = {};
    for (var i = 0; i < values.length; i++) {
        var v = values[i];
        if (v && !seen[v]) {
            seen[v] = true;
            unique.push(v);
        }
    }
    unique.sort();
    for (var i = 0; i < unique.length; i++) {
        var opt = document.createElement('option');
        opt.value = unique[i];
        opt.textContent = unique[i];
        select.appendChild(opt);
    }
}

// ============================================================
// INITIALIZATION
// ============================================================
document.addEventListener('DOMContentLoaded', function() {
    populateDropdown('filter-library', ALL_DATA.map(function(r) { return r.library; }));
    populateDropdown('filter-class-type', ALL_DATA.map(function(r) { return r.classType; }));

    var sfSelect = document.getElementById('filter-subfield');
    var noneOpt = document.createElement('option');
    noneOpt.value = 'none'; noneOpt.textContent = 'None needed';
    sfSelect.appendChild(noneOpt);
    var anyOpt = document.createElement('option');
    anyOpt.value = 'any'; anyOpt.textContent = 'Any change';
    sfSelect.appendChild(anyOpt);
    var sfTypes = [];
    var sfSeen = {};
    ALL_DATA.forEach(function(r) {
        if (r.subfieldChanges) {
            r.subfieldChanges.split(';').forEach(function(s) {
                var trimmed = s.trim();
                if (trimmed && !sfSeen[trimmed]) {
                    sfSeen[trimmed] = true;
                    sfTypes.push(trimmed);
                }
            });
        }
    });
    sfTypes.sort();
    sfTypes.forEach(function(t) {
        var opt = document.createElement('option');
        opt.value = t; opt.textContent = t;
        sfSelect.appendChild(opt);
    });

    // Attach filter listeners
    document.getElementById('filter-library').addEventListener('change', applyFilters);
    document.getElementById('filter-change').addEventListener('change', applyFilters);
    document.getElementById('filter-confidence').addEventListener('change', applyFilters);
    document.getElementById('filter-class-type').addEventListener('change', applyFilters);
    document.getElementById('filter-subfield').addEventListener('change', applyFilters);
    document.getElementById('filter-suppressed').addEventListener('change', applyFilters);
    document.getElementById('search-box').addEventListener('input', applyFilters);

    // Show default sort arrows
    var ths = document.querySelectorAll('#data-table th .sort-arrow');
    for (var i = 0; i < ths.length; i++) {
        ths[i].textContent = '\\u25E5';
    }

    // Initial filter + render
    applyFilters();
});
</script>
</body>
</html>"""


def create_html_report(df, output_path):
    """Create a self-contained interactive HTML report with analysis results.

    Shows only records where Change Needed is 'Yes' or 'Review' by default,
    with filters to expand the view. Includes a summary dashboard, sortable
    table, text search, and CSV export.

    Args:
        df: DataFrame with all analysis columns (same as create_excel_output).
        output_path: Path for the HTML file.
    """

    # Compute dashboard statistics from the FULL dataset before filtering
    total = len(df)
    count_yes = int((df.get('Change Needed') == 'Yes').sum()) if 'Change Needed' in df.columns else 0
    count_review = int((df.get('Change Needed') == 'Review').sum()) if 'Change Needed' in df.columns else 0
    count_no = int((df.get('Change Needed') == 'No').sum()) if 'Change Needed' in df.columns else 0
    libraries = sorted(df['Library Name'].dropna().unique().tolist()) \
                if 'Library Name' in df.columns else []

    # Filter to only records needing changes — the HTML report is for
    # cataloger review, not the full dataset. The Excel has everything.
    df_changes = df[df['Change Needed'].isin(['Yes', 'Review'])].copy() \
                 if 'Change Needed' in df.columns else df.copy()

    # Build JSON data array from the filtered DataFrame
    records = []
    for _, row in df_changes.iterrows():
        records.append({
            'library': str(row.get('Library Name', '')),
            'callNumber': str(row['Extracted Call Number'])
                          if pd.notna(row.get('Extracted Call Number')) else '',
            'currentInd': str(row.get('Current Indicator', '')),
            'suggestedInd': str(row.get('Suggested Indicator', '')),
            'changeNeeded': str(row.get('Change Needed', '')),
            'classType': str(row.get('Classification Type', '')),
            'confidence': str(row.get('Confidence', '')),
            'subfieldChanges': str(row['Subfield Changes'])
                               if pd.notna(row.get('Subfield Changes')) else '',
            'notes': str(row['Notes'])
                     if pd.notna(row.get('Notes')) else '',
            'mmsId': str(row.get('MMS Id', '')),
            'holdingsId': str(row.get('Holdings ID', '')),
            'suppressed': str(row.get('Suppressed', '')),
            'primoCode': CUNY_PRIMO_CODES.get(str(row.get('Institution Name', '')), ''),
            'scopeCode': PRIMO_SCOPE_CODES.get(
                CUNY_PRIMO_CODES.get(str(row.get('Institution Name', '')), ''), ''),
        })

    json_data = json.dumps(records, ensure_ascii=False)

    gen_date = datetime.now().strftime('%B %d, %Y at %I:%M %p')

    # Replace placeholder tokens in the HTML template
    html = _HTML_TEMPLATE
    html = html.replace('__JSON_DATA__', json_data)
    html = html.replace('__DATE__', gen_date)
    html = html.replace('__TOTAL__', f'{total:,}')
    html = html.replace('__COUNT_YES__', f'{count_yes:,}')
    html = html.replace('__COUNT_REVIEW__', f'{count_review:,}')
    html = html.replace('__COUNT_NO__', f'{count_no:,}')
    html = html.replace('__COUNT_CHANGES__', f'{len(records):,}')
    html = html.replace('__COUNT_LIBRARIES__', str(len(libraries)))

    Path(output_path).write_text(html, encoding='utf-8')
    print(f"HTML report saved: {output_path}")


# =============================================================================
# MAIN
# =============================================================================

def main(input_path, output_path):
    """Main processing function."""

    print(f"Loading {input_path}...")

    # Detect input format: pull script output (headers in row 1) vs
    # Alma Analytics export (3 header rows to skip, 6 columns).
    # Try reading with headers first; fall back to Analytics format.
    df_peek = pd.read_excel(input_path, nrows=3, header=None)
    first_cell = str(df_peek.iloc[0, 0]) if not pd.isna(df_peek.iloc[0, 0]) else ''

    if first_cell in ('Permanent Call Number', 'MMS Id', '852 MARC',
                       'Holdings ID', 'Institution Name', 'Suppressed'):
        # Pull script output or clean format — headers in row 1
        df = pd.read_excel(input_path)
    else:
        # Alma Analytics export — skip header rows and assign columns
        df = pd.read_excel(input_path, skiprows=2, header=None)
        df.columns = [
            'Permanent Call Number', 'Permanent Call Number Type', '852 MARC',
            'Normalized Call Number', 'Institution Name', 'MMS Id'
        ]
        df = df.iloc[1:].reset_index(drop=True)

    # Ensure expected columns exist (fill missing ones with empty strings)
    for col in ['Permanent Call Number', 'Permanent Call Number Type', '852 MARC',
                'Normalized Call Number', 'Institution Name', 'Library Name',
                'MMS Id', 'Holdings ID', 'Suppressed']:
        if col not in df.columns:
            df[col] = ''

    # Clean up numeric institution codes. Analytics sometimes returns the
    # Alma IZ code (e.g., "6129") instead of the institution name.
    # Map known codes to names using the ALMA_IZ_CODES lookup table.
    # This affects both Institution Name and Library Name columns.
    for col in ['Institution Name', 'Library Name']:
        if col in df.columns:
            code_mask = df[col].astype(str).isin(ALMA_IZ_CODES)
            if code_mask.any():
                count = code_mask.sum()
                df.loc[code_mask, col] = (
                    df.loc[code_mask, col]
                    .astype(str)
                    .map(ALMA_IZ_CODES)
                )
                print(f"  Mapped {count} IZ codes to names in {col}")

    # Normalize Suppressed column. Analytics returns 0/1 instead of Yes/No.
    if 'Suppressed' in df.columns:
        df['Suppressed'] = df['Suppressed'].astype(str).map(
            {'1': 'Yes', '0': 'No', '1.0': 'Yes', '0.0': 'No'}
        ).fillna(df['Suppressed'])

    print(f"Loaded {len(df)} records")

    # Parse 852 MARC and extract call numbers
    print("Parsing 852 MARC fields...")
    df['Parsed MARC'] = df['852 MARC'].apply(parse_852_marc)
    marc_results = df['Parsed MARC'].apply(get_call_number_from_marc)
    df[['Extracted Call Number', 'From $j', 'J Combined', 'J Conflict']] = pd.DataFrame(
        marc_results.tolist(), index=df.index
    )

    # Check if the 852 field has any call number subfields ($h, $i, $j)
    def _has_cn_subfields(parsed):
        if not parsed:
            return False
        sf = parsed.get('subfields', {})
        return bool(sf.get('h') or sf.get('i') or sf.get('j'))

    df['Has CN Subfields'] = df['Parsed MARC'].apply(_has_cn_subfields)

    # Drop records with no call number subfields ($h, $i, $j).
    # These are holdings with only $a/$b/$c — nothing to analyze.
    no_cn_count = (~df['Has CN Subfields']).sum()
    if no_cn_count:
        df = df[df['Has CN Subfields']].reset_index(drop=True)
        print(f"  Excluded {no_cn_count} records with no call number subfields")

    df['Call Number for Analysis'] = df.apply(
        lambda row: row['Extracted Call Number'] if row['Extracted Call Number'] else row['Permanent Call Number'],
        axis=1
    )

    # Extract current indicator from parsed MARC
    # Use 'blank' for empty/missing indicators so they display clearly
    df['Current Indicator'] = df['Parsed MARC'].apply(
        lambda x: x.get('indicator1', '') if x else '').apply(
        lambda v: v if v else 'blank')

    # Classify each call number
    print("Classifying call numbers...")
    indicators = []
    class_types = []
    confidences = []
    notes = []
    subfield_changes = []

    for _, row in df.iterrows():
        cn = row['Call Number for Analysis']
        from_j = row['From $j']
        j_combined = row['J Combined']
        j_conflict = row['J Conflict']
        institution = row.get('Institution Name')
        result = categorize_call_number(cn, from_j=from_j, j_combined=j_combined,
                                        j_conflict=j_conflict, institution=institution)
        indicators.append(result[0])
        class_types.append(result[1])
        confidences.append(result[2])
        notes.append(result[3])

        # Check for trailing period on $h (misplaced cutter period)
        sf_changes = result[4]
        parsed = row.get('Parsed MARC')
        if parsed:
            sf = parsed.get('subfields', {})
            raw_h = sf.get('h', '')
            if raw_h.endswith('.') and sf.get('i', ''):
                period_note = 'Move period from end of $h to start of $i'
                if sf_changes:
                    sf_changes = sf_changes + '; ' + period_note
                else:
                    sf_changes = period_note
        subfield_changes.append(sf_changes)

    df['Suggested Indicator'] = indicators
    df['Classification Type'] = class_types
    df['Confidence'] = confidences
    df['Notes'] = notes
    df['Subfield Changes'] = subfield_changes

    # Compare current vs suggested indicator
    df['Change Needed'] = df.apply(
        lambda row: 'Yes' if row['Current Indicator'] != row['Suggested Indicator']
                          and row['Suggested Indicator'] not in ['blank', 'N/A']
                     else ('Review' if row['Suggested Indicator'] in ['N/A']
                           else 'No'),
        axis=1)

    # Print summary
    print("\nClassification Summary:")
    print("=" * 60)
    summary = df.groupby(['Suggested Indicator', 'Classification Type']).size().reset_index(name='Count')
    summary = summary.sort_values('Count', ascending=False)
    print(summary.to_string(index=False))
    
    # Create Excel output
    print(f"\nSaving to {output_path}...")
    create_excel_output(df, output_path)

    # Create interactive HTML report alongside the Excel file
    html_path = str(Path(output_path).with_suffix('.html'))
    print(f"Saving HTML report to {html_path}...")
    create_html_report(df, html_path)

    print("Done!")


if __name__ == '__main__':
    if len(sys.argv) < 2 or len(sys.argv) > 3:
        print(f"Usage: {sys.argv[0]} input.xlsx [output.xlsx]")
        sys.exit(1)

    input_file = sys.argv[1]

    if len(sys.argv) == 3:
        output_file = sys.argv[2]
    else:
        # Auto-generate timestamped output name from input name
        # e.g., KB_852_data_20260301_120236.xlsx → KB_852_analyzed_20260301_120236_HHMMSS.xlsx
        stem = Path(input_file).stem
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        # Replace "_data" with "_analyzed" if present, otherwise append "_analyzed"
        if '_data' in stem:
            out_stem = stem.replace('_data', '_analyzed', 1)
        else:
            out_stem = f"{stem}_analyzed"
        output_file = str(Path(input_file).parent / f"{out_stem}_{timestamp}.xlsx")

    main(input_file, output_file)
