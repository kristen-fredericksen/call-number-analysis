#!/usr/bin/env python3
"""
Pull 852 field data from Alma Analytics via the REST API.

Retrieves the "852 Field Analysis - All Indicators" report for one or more
CUNY schools and saves the results as an Excel file compatible with
analyze_852_indicators.py.

Usage:
    python pull_852_analytics.py KB              # Pull Kingsborough only
    python pull_852_analytics.py KB BM           # Pull Kingsborough and BMCC
    python pull_852_analytics.py --all           # Pull all schools with keys

Output: data/{school_code}_852_data_{timestamp}.xlsx (one file per school)

Requires: api_keys.env in the project root with IZ API keys.
"""

import os
import sys
import time
from datetime import datetime
import requests
import xml.etree.ElementTree as ET
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

# Project root (one level up from src/)
PROJECT_ROOT = Path(__file__).parent.parent

# Analytics report paths per school. Add new schools here as reports are created.
REPORT_PATHS = {
    'BM': '/shared/Manhattan Community College 01CUNY_BM/Reports/852 Field Analysis - All Indicators',
    'KB': '/shared/Kingsborough Community College 01CUNY_KB/Cataloging/852 Field Analysis - All Indicators',
    'QC': '/shared/Queens College 01CUNY_QC/Reports/852 Field Analysis - All Indicators',
}

# Map school codes to their API key variable names in api_keys.env
SCHOOL_KEY_MAP = {
    'BB': 'BB_IZ_API_KEY',
    'BM': 'BM_IZ_API_KEY',
    'BX': 'BX_IZ_API_KEY',
    'BC': 'BC_IZ_API_KEY',
    'SI': 'SI_IZ_API_KEY',
    'GJ': 'GJ_IZ_API_KEY',
    'AL': 'AL_IZ_API_KEY',
    'GC': 'GC_IZ_API_KEY',
    'CL': 'CL_IZ_API_KEY',
    'NC': 'NC_IZ_API_KEY',
    'HO': 'HO_IZ_API_KEY',
    'HC': 'HC_IZ_API_KEY',
    'JJ': 'JJ_IZ_API_KEY',
    'KB': 'KB_IZ_API_KEY',
    'LG': 'LG_IZ_API_KEY',
    'LE': 'LE_IZ_API_KEY',
    'ME': 'ME_IZ_API_KEY',
    'NY': 'NY_IZ_API_KEY',
    'QC': 'QC_IZ_API_KEY',
    'QB': 'QB_IZ_API_KEY',
    'CC': 'CC_IZ_API_KEY',
    'YC': 'YC_IZ_API_KEY',
}

# Expected columns from the Analytics report (in the order Analytics returns
# them, which is alphabetical by full path). The dummy "Column 0" is skipped.
# These get mapped to the column names the analysis script expects.
ANALYTICS_TO_SCRIPT_COLUMNS = {
    'MMS Id': 'MMS Id',
    '852 MARC': '852 MARC',
    'Holdings ID': 'Holdings ID',
    'Normalized Call Number': 'Normalized Call Number',
    'Permanent Call Number Type': 'Permanent Call Number Type',
    'Permanent Call Number': 'Permanent Call Number',
    'Suppressed from Discovery': 'Suppressed',
    'Institution Name': 'Institution Name',
    'Library Name': 'Library Name',
}


def load_env(env_path):
    """Load key=value pairs from an env file."""
    config = {}
    if not env_path.exists():
        return config
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            if '=' in line:
                key, value = line.split('=', 1)
                config[key.strip()] = value.strip()
    return config


def get_api_key(config, school_code):
    """Get the API key for a school from the loaded config."""
    key_name = SCHOOL_KEY_MAP.get(school_code)
    if not key_name:
        return None
    key = config.get(key_name, '')
    return key if key else None


def fetch_analytics_report(base_url, api_key, report_path, limit=1000):
    """
    Fetch an Alma Analytics report with pagination.

    Returns a list of dicts, one per row, with column names as keys.
    """
    url = f"{base_url}/almaws/v1/analytics/reports"
    all_rows = []
    column_names = None
    token = None
    page = 0

    while True:
        page += 1
        params = {
            'apikey': api_key,
            'limit': limit,
            'col_names': 'true',
        }

        if token:
            params['token'] = token
        else:
            params['path'] = report_path

        print(f"  Fetching page {page}...", end=' ', flush=True)
        response = requests.get(url, params=params)

        if response.status_code != 200:
            print(f"ERROR (HTTP {response.status_code})")
            print(f"  Response: {response.text[:500]}")
            sys.exit(1)

        root = ET.fromstring(response.text)

        # Check if finished
        is_finished_el = root.find('.//IsFinished')
        is_finished = is_finished_el is not None and is_finished_el.text == 'true'

        # Find the result XML — it may be nested inside a namespace
        result_xml = root.find('.//ResultXml')
        if result_xml is None:
            print("ERROR: No ResultXml in response")
            print(f"  Response: {response.text[:500]}")
            sys.exit(1)

        # The actual data is inside a rowset element (may have namespace)
        # Find all elements regardless of namespace
        rowset = None
        for elem in result_xml.iter():
            if 'rowset' in elem.tag.lower():
                rowset = elem
                break

        if rowset is None:
            # Try without namespace
            rowset = result_xml

        # Extract column names from the first page
        if column_names is None:
            # Column names come from the schema or the first Row's structure
            # In Analytics API, column headers are in a separate element
            # or we extract them from the xsd:schema
            column_names = _extract_column_names(root)
            if column_names:
                print(f"({len(column_names)} columns)", end=' ')

        # Extract rows. Column elements have tags like "Column0", "Column5",
        # etc. When a column is empty, Analytics may skip it entirely, so we
        # must parse the column index from the tag and place each value at
        # the correct position (not just append in order).
        rows_found = 0
        for row_elem in rowset.iter():
            if 'Row' in row_elem.tag and row_elem.tag != rowset.tag:
                indexed_values = {}
                for col_elem in row_elem:
                    # Tag might be namespaced: {ns}Column5 → extract "5"
                    tag = col_elem.tag.split('}')[-1] if '}' in col_elem.tag else col_elem.tag
                    # Extract the column number from the tag (e.g., "Column5" → 5)
                    col_num = ''.join(c for c in tag if c.isdigit())
                    if col_num:
                        indexed_values[int(col_num)] = col_elem.text if col_elem.text else ''
                if indexed_values:
                    all_rows.append(indexed_values)
                    rows_found += 1

        print(f"{rows_found} rows")

        if is_finished:
            break

        # Get resumption token
        token_el = root.find('.//ResumptionToken')
        if token_el is None or not token_el.text:
            print("  Warning: No resumption token but IsFinished is false. Stopping.")
            break
        token = token_el.text

        # Brief pause to be polite to the API
        time.sleep(0.5)

    return column_names, all_rows


def _extract_column_names(root):
    """
    Extract column names from the Analytics API response.

    Returns a dict mapping column index (int) to column heading (str),
    e.g. {0: 'Column 0', 5: 'MMS Id', 6: '852 MARC', ...}.

    The schema elements are named like "Column0", "Column5", etc., and
    each has a columnHeading attribute with the human-readable name.
    Matching by index (not list position) is essential because different
    schools' reports may have different extra columns, and Analytics
    skips empty columns in the XML rows.
    """
    indexed_names = {}

    # Look for columnHeading attributes in the schema, paired with the
    # Column index from the element's 'name' attribute (e.g., "Column5").
    for elem in root.iter():
        heading = elem.attrib.get('{urn:saw-sql}columnHeading')
        if not heading:
            for attr_name, attr_val in elem.attrib.items():
                if 'columnHeading' in attr_name:
                    heading = attr_val
                    break
        if heading:
            # Get the column index from the 'name' attribute (e.g., "Column5")
            col_tag = elem.attrib.get('name', '')
            col_num = ''.join(c for c in col_tag if c.isdigit())
            if col_num:
                indexed_names[int(col_num)] = heading

    return indexed_names if indexed_names else None


def rows_to_dataframe(column_names_map, rows):
    """
    Convert raw Analytics rows to a DataFrame with the columns the
    analysis script expects.

    column_names_map is a dict mapping column index → heading name
    (from the schema), e.g. {0: 'Column 0', 5: 'MMS Id', ...}.
    Each row is also a dict mapping column index → value (from the
    XML parsing). Both are matched by index so column order and
    missing columns don't matter.
    """
    if not column_names_map or not rows:
        print("  Warning: No column names or no data rows.")
        return pd.DataFrame()

    # Build a list of (index, analytics_name, script_name) for columns
    # we want to keep — i.e., columns whose Analytics heading matches
    # something in ANALYTICS_TO_SCRIPT_COLUMNS.
    keep_columns = []
    for idx, heading in sorted(column_names_map.items()):
        # Try exact match first
        if heading in ANALYTICS_TO_SCRIPT_COLUMNS:
            keep_columns.append((idx, ANALYTICS_TO_SCRIPT_COLUMNS[heading]))
            continue
        # Fall back to substring match
        for analytics_name, script_name in ANALYTICS_TO_SCRIPT_COLUMNS.items():
            if analytics_name in heading:
                keep_columns.append((idx, script_name))
                break

    if not keep_columns:
        print("  Warning: No matching columns found in Analytics response.")
        return pd.DataFrame()

    # Build the DataFrame by pulling the correct index from each row.
    data = []
    for row in rows:
        data.append([row.get(idx, '') for idx, _ in keep_columns])

    col_names = [name for _, name in keep_columns]
    df = pd.DataFrame(data, columns=col_names)

    return df


def save_to_excel(df, output_path):
    """Save DataFrame to Excel with 12pt Arial formatting."""
    data_font = Font(name='Arial', size=12)
    header_font = Font(name='Arial', size=12, bold=True)

    # Force ID columns to string so Excel doesn't convert them to
    # scientific notation (which loses digits on long numbers).
    id_columns = {'MMS Id', 'Holdings ID'}
    for col in id_columns:
        if col in df.columns:
            df[col] = df[col].astype(str)

    wb = Workbook()
    ws = wb.active
    ws.title = "Analytics Data"

    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font

    # Write data
    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row, 1):
            col_name = df.columns[col_idx - 1]
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
            cell.font = data_font
            # Store IDs as explicit text to prevent scientific notation
            if col_name in id_columns:
                cell.number_format = '@'

    # Auto-width columns
    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = chr(64 + col_idx) if col_idx <= 26 else 'A'
        ws.column_dimensions[col_letter].width = max(15, len(str(col_name)) + 5)

    wb.save(output_path)


def main():
    if len(sys.argv) < 2:
        print("Usage: python pull_852_analytics.py <school_code> [school_code ...]")
        print("       python pull_852_analytics.py --all")
        print()
        print("School codes: BB, BM, BX, BC, SI, GJ, AL, GC, CL, NC, HO,")
        print("              HC, JJ, KB, LG, LE, ME, NY, QC, QB, CC, YC")
        sys.exit(1)

    # Load API keys
    env_path = PROJECT_ROOT / 'api_keys.env'
    config = load_env(env_path)
    if not config:
        print(f"Error: Could not load {env_path}")
        print("Make sure api_keys.env exists in the project root.")
        sys.exit(1)

    base_url = config.get('ALMA_API_BASE_URL', 'https://api-na.hosted.exlibrisgroup.com')

    # Determine which schools to pull
    if '--all' in sys.argv:
        schools = [code for code in SCHOOL_KEY_MAP if get_api_key(config, code)]
    else:
        schools = [code.upper() for code in sys.argv[1:]]

    # Validate
    for code in schools:
        if code not in SCHOOL_KEY_MAP:
            print(f"Error: Unknown school code '{code}'")
            sys.exit(1)
        if not get_api_key(config, code):
            print(f"Warning: No API key for {code}. Skipping.")
            schools = [s for s in schools if s != code]
        if code not in REPORT_PATHS:
            print(f"Error: No Analytics report path configured for {code}.")
            print(f"Add the path to REPORT_PATHS in this script.")
            sys.exit(1)

    if not schools:
        print("No schools to process.")
        sys.exit(1)

    # Process each school
    data_dir = PROJECT_ROOT / 'data'
    data_dir.mkdir(exist_ok=True)

    for code in schools:
        api_key = get_api_key(config, code)
        report_path = REPORT_PATHS[code]

        print(f"\n{'='*60}")
        print(f"Pulling data for {code}")
        print(f"Report: {report_path}")
        print(f"{'='*60}")

        column_names, rows = fetch_analytics_report(base_url, api_key, report_path)

        if not rows:
            print(f"  No data returned for {code}.")
            continue

        print(f"\n  Total rows: {len(rows)}")

        # Convert to DataFrame
        df = rows_to_dataframe(column_names, rows)
        print(f"  Columns: {list(df.columns)}")

        # Save
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = data_dir / f"{code}_852_data_{timestamp}.xlsx"
        save_to_excel(df, output_path)
        print(f"  Saved to: {output_path}")

    print("\nDone!")


if __name__ == '__main__':
    main()
